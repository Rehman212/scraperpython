"""Resumable exhaustive UPrinting price-matrix crawler."""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import sqlite3
import threading
import time
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from pathlib import Path
from typing import Any, Iterator

import requests
from openpyxl import Workbook

from uprinting_scraper import DEFAULT_URL, UPrintingScraper

LOG = logging.getLogger("full-scrape")
THREAD = threading.local()


class TransientScrapeError(RuntimeError):
    pass


def key_for(selection: dict[str, str]) -> str:
    raw = "&".join(f"{k}={selection[k]}" for k in sorted(selection))
    return hashlib.sha1(raw.encode()).hexdigest()


class MatrixCrawler:
    def __init__(self, url: str, database: Path, workers: int, delay: float) -> None:
        self.scraper = UPrintingScraper(url)
        self.scraper.load()
        self.database = database
        self.workers = workers
        self.delay = delay
        self.attributes = [a for a in self.scraper.attributes() if a["attribute_id"] != "3" or any(o["option_id"] != "custom" for o in a["options"])]
        self.raw_attrs = self.scraper.catalog["prod_attrs"]
        self.size_rules = self.scraper.catalog.get("new_product_size_exceptions", {})
        self.option_lookup = {a["attribute_id"]: {o["option_id"]: o for o in a["options"]} for a in self.attributes}
        self._init_db()

    def _connect(self) -> sqlite3.Connection:
        db = sqlite3.connect(self.database, timeout=60)
        db.execute("PRAGMA journal_mode=WAL")
        db.execute("PRAGMA synchronous=NORMAL")
        return db

    def _init_db(self) -> None:
        with self._connect() as db:
            db.executescript("""
                CREATE TABLE IF NOT EXISTS prices(
                  config_key TEXT PRIMARY KEY, selection TEXT NOT NULL, labels TEXT,
                  price REAL, unit_price REAL, quantity INTEGER, turnaround_days INTEGER,
                  in_stock TEXT, scraped_at INTEGER NOT NULL
                );
                CREATE TABLE IF NOT EXISTS errors(
                  config_key TEXT PRIMARY KEY, selection TEXT NOT NULL, error TEXT NOT NULL,
                  scraped_at INTEGER NOT NULL
                );
            """)

    @staticmethod
    def _matches(rule: dict[str, Any], selection: dict[str, str]) -> bool:
        return all(selection.get(f"attr{k}") == str(v) for k, v in rule.items())

    def _within_size(self, attr_id: str, option_id: str, selection: dict[str, str]) -> bool:
        rules = self.size_rules.get(attr_id, {}).get(option_id)
        if not rules:
            return True
        size_id = selection.get("attr3")
        size = self.option_lookup.get("3", {}).get(size_id or "", {}).get("factors", {})
        if not size or size_id == "custom":
            return False
        width, height = float(size.get("width", 0)), float(size.get("height", 0))
        for rule in rules:
            minimum_w, maximum_w = float(rule.get("min_width") or 0), float(rule.get("max_width") or 0)
            minimum_h, maximum_h = float(rule.get("min_height") or 0), float(rule.get("max_height") or 0)
            w_ok = (not minimum_w or width >= minimum_w) and (not maximum_w or width <= maximum_w)
            h_ok = (not minimum_h or height >= minimum_h) and (not maximum_h or height <= maximum_h)
            if (w_ok and h_ok) if rule.get("condition", "and") == "and" else (w_ok or h_ok):
                return True
        return False

    def selections(self) -> Iterator[dict[str, str]]:
        def walk(index: int, selected: dict[str, str]) -> Iterator[dict[str, str]]:
            if index == len(self.attributes):
                yield dict(selected)
                return
            attr = self.attributes[index]
            attr_id = attr["attribute_id"]
            exceptions = self.raw_attrs[attr_id].get("exceptions", {})
            hidden = any(self._matches(rule, selected) for rule in exceptions.get("-1", []))
            if hidden:
                candidates = [next((o for o in attr["options"] if o["option_id"] == attr["default_option_id"]), attr["options"][0])]
            else:
                candidates = [
                    option for option in attr["options"]
                    if option["option_id"] != "custom"
                    and not any(self._matches(rule, selected) for rule in exceptions.get(option["option_id"], []))
                    and self._within_size(attr_id, option["option_id"], selected)
                ]
            for option in candidates:
                selected[f"attr{attr_id}"] = option["option_id"]
                yield from walk(index + 1, selected)
            selected.pop(f"attr{attr_id}", None)
        yield from walk(0, {})

    def _fetch(self, selection: dict[str, str]) -> dict[str, Any]:
        if not hasattr(THREAD, "scraper"):
            worker = UPrintingScraper(self.scraper.url)
            worker.product_id = self.scraper.product_id
            worker.api_url = self.scraper.api_url
            worker.auth = self.scraper.auth
            THREAD.scraper = worker
        last_error: Exception | None = None
        for attempt in range(8):
            if self.delay:
                time.sleep(self.delay)
            try:
                response = THREAD.scraper.price(selection)
                break
            except Exception as exc:
                last_error = exc
                message = str(exc)
                if not any(code in message for code in ("API 403", "API 429", "API 500", "API 502", "API 503", "API 504")):
                    raise
                time.sleep(min(120, 5 * (2**attempt)))
        else:
            raise TransientScrapeError(str(last_error))
        labels = {str(x["attribute_id"]): x.get("attr_value", "") for x in response.get("display_specs", [])}
        return {
            "labels": labels, "price": response.get("price"), "unit_price": response.get("unit_price"),
            "quantity": response.get("qty"), "turnaround_days": response.get("turnaround"),
            "in_stock": response.get("in_stock_flag"),
        }

    def run(self, status_path: Path) -> None:
        with self._connect() as db:
            completed = {row[0] for row in db.execute("SELECT config_key FROM prices UNION SELECT config_key FROM errors")}
        LOG.info("Resume checkpoint: %s completed configurations", f"{len(completed):,}")
        submitted = done = valid = invalid = transient_retries = 0
        started = time.time()
        iterator = ((key_for(s), s) for s in self.selections())
        pending: dict[Any, tuple[str, dict[str, str]]] = {}
        max_pending = self.workers * 4

        def update_status() -> None:
            elapsed = max(time.time() - started, 0.001)
            status_path.write_text(json.dumps({
                "status": "running", "already_complete": len(completed), "submitted_this_run": submitted,
                "finished_this_run": done, "valid_this_run": valid, "invalid_this_run": invalid,
                "transient_retries": transient_retries,
                "requests_per_second": round(done / elapsed, 2), "updated_at": int(time.time()),
            }, indent=2), encoding="utf-8")

        with self._connect() as db, ThreadPoolExecutor(max_workers=self.workers) as pool:
            exhausted = False
            while pending or not exhausted:
                while len(pending) < max_pending and not exhausted:
                    try:
                        config_key, selection = next(iterator)
                    except StopIteration:
                        exhausted = True
                        break
                    if config_key in completed:
                        continue
                    future = pool.submit(self._fetch, selection)
                    pending[future] = (config_key, selection)
                    submitted += 1
                if not pending:
                    continue
                finished, _ = wait(pending, return_when=FIRST_COMPLETED)
                for future in finished:
                    config_key, selection = pending.pop(future)
                    now = int(time.time())
                    try:
                        row = future.result()
                        db.execute(
                            "INSERT OR REPLACE INTO prices VALUES(?,?,?,?,?,?,?,?,?)",
                            (config_key, json.dumps(selection, separators=(",", ":")), json.dumps(row["labels"], separators=(",", ":")),
                             row["price"], row["unit_price"], row["quantity"], row["turnaround_days"], row["in_stock"], now),
                        )
                        valid += 1
                    except TransientScrapeError:
                        # WAF/rate-limit failures are never permanent invalid rows.
                        # Put the configuration back into the queue after cooldown.
                        transient_retries += 1
                        update_status()
                        LOG.warning("Temporary UPrinting block; cooling down before retry %s", config_key)
                        time.sleep(30)
                        retry = pool.submit(self._fetch, selection)
                        pending[retry] = (config_key, selection)
                        continue
                    except Exception as exc:
                        db.execute("INSERT OR REPLACE INTO errors VALUES(?,?,?,?)", (config_key, json.dumps(selection), str(exc), now))
                        invalid += 1
                    done += 1
                    if done % 100 == 0:
                        db.commit(); update_status()
                    if done % 1000 == 0:
                        LOG.info("Finished %s | valid %s | invalid %s", f"{done:,}", f"{valid:,}", f"{invalid:,}")
            db.commit()
        status_path.write_text(json.dumps({"status": "complete", "finished_this_run": done, "valid_this_run": valid, "invalid_this_run": invalid}, indent=2), encoding="utf-8")

    def export(self, json_path: Path, xlsx_path: Path) -> None:
        attr_ids = [a["attribute_id"] for a in self.attributes]
        attr_names = {a["attribute_id"]: a["name"] for a in self.attributes}
        with self._connect() as db, json_path.open("w", encoding="utf-8") as stream:
            valid_count = db.execute("SELECT COUNT(*) FROM prices").fetchone()[0]
            error_count = db.execute("SELECT COUNT(*) FROM errors").fetchone()[0]
            metadata = {
                "source_url": self.scraper.url, "product_id": self.scraper.product_id,
                "product_name": self.scraper.catalog.get("product_name"), "mode": "dependency_pruned_exhaustive",
                "valid_price_rows": valid_count, "invalid_rows": error_count,
                "custom_size_note": "Custom dimensions are continuous inputs and are not expanded into the finite matrix.",
            }
            stream.write('{"metadata":')
            json.dump(metadata, stream, separators=(",", ":"))
            stream.write(',"attributes":')
            json.dump(self.attributes, stream, separators=(",", ":"), ensure_ascii=False)
            stream.write(',"prices":[')
            first = True
            for selection, labels, price, unit, qty, turnaround, stock in db.execute(
                "SELECT selection,labels,price,unit_price,quantity,turnaround_days,in_stock FROM prices ORDER BY rowid"
            ):
                if not first:
                    stream.write(",")
                first = False
                json.dump({"selection": json.loads(selection), "display": json.loads(labels), "price": price,
                           "unit_price": unit, "quantity": qty, "turnaround_days": turnaround,
                           "in_stock": stock, "currency": "USD"}, stream, separators=(",", ":"), ensure_ascii=False)
            stream.write('],"errors":[')
            first = True
            for selection, error in db.execute("SELECT selection,error FROM errors ORDER BY rowid"):
                if not first:
                    stream.write(",")
                first = False
                json.dump({"selection": json.loads(selection), "error": error}, stream, separators=(",", ":"), ensure_ascii=False)
            stream.write("]}")

        workbook = Workbook(write_only=True)
        headers = ["price", "unit_price", "quantity", "turnaround_days", "in_stock"]
        for attr_id in attr_ids:
            headers.extend([f"{attr_names[attr_id]} ID", attr_names[attr_id]])
        sheet = None
        row_in_sheet = 0
        with self._connect() as db:
            for selection_raw, labels_raw, price, unit, qty, turnaround, stock in db.execute(
                "SELECT selection,labels,price,unit_price,quantity,turnaround_days,in_stock FROM prices ORDER BY rowid"
            ):
                if sheet is None or row_in_sheet >= 1_000_000:
                    sheet = workbook.create_sheet(f"Prices {len(workbook.worksheets) + 1}")
                    sheet.append(headers)
                    row_in_sheet = 0
                selection, labels = json.loads(selection_raw), json.loads(labels_raw)
                row: list[Any] = [price, unit, qty, turnaround, stock]
                for attr_id in attr_ids:
                    row.extend([selection.get(f"attr{attr_id}", ""), labels.get(attr_id, "")])
                sheet.append(row)
                row_in_sheet += 1
            errors = workbook.create_sheet("Errors")
            errors.append(["selection", "error"])
            for selection, error in db.execute("SELECT selection,error FROM errors ORDER BY rowid"):
                errors.append([selection, error])
        workbook.save(xlsx_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("url", nargs="?", default=DEFAULT_URL)
    parser.add_argument("--database", default="full_prices.sqlite")
    parser.add_argument("--status", default="full_scrape_status.json")
    parser.add_argument("--workers", type=int, default=1)
    parser.add_argument("--delay", type=float, default=1.5)
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    status_path = Path(args.status)
    while True:
        try:
            crawler = MatrixCrawler(args.url, Path(args.database), max(1, args.workers), max(0, args.delay))
            break
        except requests.HTTPError as exc:
            status = exc.response.status_code if exc.response is not None else None
            if status not in (403, 429):
                raise
            status_path.write_text(json.dumps({
                "status": "waiting_for_uprinting",
                "reason": f"HTTP {status} temporary access block",
                "retry_after_seconds": 300,
                "updated_at": int(time.time()),
            }, indent=2), encoding="utf-8")
            LOG.warning("UPrinting returned HTTP %s; retrying product page in 5 minutes", status)
            time.sleep(300)
    crawler.run(status_path)
    base = Path(args.database).with_suffix("")
    crawler.export(Path(str(base) + ".printoe.json"), base.with_suffix(".xlsx"))
