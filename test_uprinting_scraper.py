from pathlib import Path

from openpyxl import load_workbook

from uprinting_scraper import UPrintingScraper, build_export, save_json, save_xlsx
import json

import pytest

import preview_server
from preview_server import (
    SCRAPER as PREVIEW_SCRAPER,
    calculator_export_attributes,
    deduplicate_prices,
    export_fingerprint,
    export_is_current,
    normalize_selection,
    validate_printoe_export,
)


URL = "https://www.uprinting.com/brochure-printing.html"


class FakeScraper:
    def __init__(self, option_id: str = "one") -> None:
        self.url = "https://www.uprinting.com/example.html"
        self.product_id = "123"
        self.defaults = {"attr1": option_id}
        self._option_id = option_id
        self.catalog = {"prod_attrs": {"1": {"exceptions": {}}}}

    def attributes(self):
        return [{
            "attribute_id": "1",
            "name": "Size",
            "default_option_id": self._option_id,
            "sort_order": 1,
            "options": [{"option_id": self._option_id, "label": self._option_id}],
        }]


def valid_export() -> dict:
    return {
        "metadata": {
            "source_url": "https://www.uprinting.com/example.html",
            "product_id": "123",
        },
        "attributes": [{"attribute_id": "1", "options": [{"option_id": "one"}]}],
        "prices": [{
            "selection": {"attr1": "one"},
            "price": 10,
            "unit_price": 1,
            "quantity": 10,
        }],
    }


def test_export_fingerprint_invalidates_changed_configuration(tmp_path: Path) -> None:
    first = FakeScraper("one")
    second = FakeScraper("two")
    first_fingerprint = export_fingerprint(first, {"123": first})
    assert first_fingerprint != export_fingerprint(second, {"123": second})

    export_path = tmp_path / "example.printoe.json"
    export_path.write_text(json.dumps({
        "metadata": {"export_fingerprint": first_fingerprint},
    }), encoding="utf-8")
    assert export_is_current(export_path, first_fingerprint)
    assert not export_is_current(export_path, "stale")


def test_export_validation_and_price_deduplication() -> None:
    data = valid_export()
    validate_printoe_export(data)
    duplicate = dict(data["prices"][0])
    duplicate["price"] = 999
    rows = deduplicate_prices([data["prices"][0], duplicate])
    assert len(rows) == 1
    assert rows[0]["price"] == 10

    data["prices"][0]["unit_price"] = None
    with pytest.raises(ValueError, match="unit_price"):
        validate_printoe_export(data)


def test_dynamic_rules_and_variant_defaults_are_exported() -> None:
    variant = FakeScraper("no")
    variant.product_id = "standard"
    variant.defaults = {"attr1": "no", "attr2": "square"}
    variant.catalog = {
        "prod_attrs": {
            "1": {"exceptions": {"yes": [{"2": "square"}]}},
            "2": {"exceptions": {}},
        },
    }
    variant.attributes = lambda: [
        {
            "attribute_id": "1",
            "name": "Rounded Corners",
            "default_option_id": "no",
            "sort_order": 1,
            "options": [
                {"option_id": "no", "label": "No", "sort_order": 1},
                {"option_id": "yes", "label": "Yes", "sort_order": 2},
            ],
        },
        {
            "attribute_id": "2",
            "name": "Shape",
            "default_option_id": "square",
            "sort_order": 2,
            "options": [{"option_id": "square", "label": "Square", "sort_order": 1}],
        },
    ]

    attributes = calculator_export_attributes({"standard": variant})
    rounded = next(item for item in attributes if item["attribute_id"] == "1")
    yes = next(item for item in rounded["options"] if item["option_id"] == "yes")
    assert rounded["defaults_by_product"] == {"standard": "no"}
    assert yes["available_product_ids"] == ["standard"]
    assert yes["exclusion_rules_by_product"]["standard"] == [{"2": "square"}]


def test_invalid_sweep_selection_is_rejected_before_pricing() -> None:
    scraper = UPrintingScraper("https://www.uprinting.com/example.html")
    scraper.catalog = {
        "prod_attrs": {
            "1": {"exceptions": {"yes": [{"2": "square"}]}},
        },
    }
    assert not scraper.selection_is_valid({"attr1": "yes", "attr2": "square"})
    assert scraper.selection_is_valid({"attr1": "no", "attr2": "square"})


def test_admin_login_requires_environment_password(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(preview_server, "_printoe_token", None)
    monkeypatch.setattr(preview_server, "PRINTOE_ADMIN_PASSWORD", "")
    with pytest.raises(RuntimeError, match="PRINTOE_ADMIN_PASSWORD"):
        preview_server.printoe_admin_token()


def test_admin_login_propagates_api_failure(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(preview_server, "_printoe_token", None)
    monkeypatch.setattr(preview_server, "PRINTOE_ADMIN_PASSWORD", "configured")

    def fail(*_args, **_kwargs):
        raise RuntimeError("Printoe API error (401): invalid credentials")

    monkeypatch.setattr(preview_server, "_printoe_request", fail)
    with pytest.raises(RuntimeError, match="401"):
        preview_server.printoe_admin_token()


def test_live_page_catalog_and_default_price(tmp_path: Path) -> None:
    scraper = UPrintingScraper(URL)
    scraper.load()
    assert scraper.product_id == "4"
    attributes = scraper.attributes()
    assert any(a["name"] == "Size (before folding)" for a in attributes)
    assert all(a["options"] for a in attributes)
    response = scraper.price(scraper.defaults)
    assert float(response["price"]) > 0
    assert response["display_specs"]

    row = scraper._priced_row(scraper.defaults)
    data = build_export(scraper, "default", [row], [])
    json_path = tmp_path / "result.json"
    xlsx_path = tmp_path / "result.xlsx"
    save_json(data, json_path)
    save_xlsx(data, xlsx_path)
    assert json_path.stat().st_size > 1000
    workbook = load_workbook(xlsx_path, read_only=True)
    assert {"Summary", "Attributes", "Prices", "Errors"}.issubset(workbook.sheetnames)
    assert workbook["Prices"].max_row == 2


def test_live_two_quantity_prices() -> None:
    scraper = UPrintingScraper(URL)
    scraper.load()
    quantity = next(a for a in scraper.attributes() if a["name"] == "Quantity")
    values = quantity["options"][:2]
    prices = []
    for option in values:
        selection = dict(scraper.defaults)
        selection[f"attr{quantity['attribute_id']}"] = option["option_id"]
        prices.append(float(scraper.price(selection)["price"]))
    assert len(prices) == 2
    assert all(price > 0 for price in prices)


def test_preview_matches_uprinting_reference_configuration() -> None:
    PREVIEW_SCRAPER.load()
    size = next(a for a in PREVIEW_SCRAPER.attributes() if a["attribute_id"] == "3")
    assert len(size["options"]) == 10
    selection = {
        "attr1": "69028", "attr3": "143", "attr4": "165", "attr5": "169",
        "attr6": "199", "attr7": "151", "attr16": "194516",
        "attr400": "68456", "attr635": "68458",
    }
    normalized = normalize_selection(selection, protected_attr="7")
    assert normalized["attr4"] == "162"  # UPrinting changes this to Front Only.
    result = PREVIEW_SCRAPER.price(normalized)
    assert result["price"] == "268.20"
    assert result["unit_price"] == "0.5364"


def test_preview_matches_large_brochure_single_and_double_side_prices() -> None:
    PREVIEW_SCRAPER.load()
    base = {
        "attr1": "69028", "attr3": "14731", "attr5": "169", "attr6": "199",
        "attr7": "69030", "attr16": "194516", "attr400": "68456", "attr635": "68458",
    }
    expected = {"164": ("Outside Only", "1089.50"), "165": ("Outside and Inside", "1939.80")}
    for side_id, (side_label, price) in expected.items():
        result = PREVIEW_SCRAPER.price({**base, "attr4": side_id})
        selected_side = next(x for x in result["display_specs"] if x["attribute_id"] == "4")
        assert selected_side["attr_value"] == side_label
        assert result["price"] == price
