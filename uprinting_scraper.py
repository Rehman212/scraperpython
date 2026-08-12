"""Export UPrinting product variation IDs and live prices to JSON/XLSX."""

from __future__ import annotations

import argparse
import base64
import itertools
import json
import logging
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

LOG = logging.getLogger("uprinting")
DEFAULT_URL = "https://www.uprinting.com/brochure-printing.html"


class ScraperError(RuntimeError):
    pass


def _first(pattern: str, text: str, description: str, flags: int = 0) -> str:
    match = re.search(pattern, text, flags)
    if not match:
        raise ScraperError(f"Page se {description} nahi mila; site markup shayad change ho gaya hai.")
    return match.group(1)


class UPrintingScraper:
    def __init__(self, url: str, timeout: float = 30, retries: int = 4) -> None:
        parsed = urlparse(url)
        if parsed.scheme not in {"http", "https"} or not parsed.hostname:
            raise ValueError("Valid http(s) product URL dein.")
        if not parsed.hostname.endswith("uprinting.com"):
            raise ValueError("Yeh scraper sirf uprinting.com URLs accept karta hai.")
        self.url = url
        self.timeout = timeout
        retry = Retry(
            total=retries,
            backoff_factor=0.7,
            status_forcelist=(429, 500, 502, 503, 504),
            allowed_methods=frozenset({"GET", "POST"}),
        )
        self.session = requests.Session()
        self.session.mount("https://", HTTPAdapter(max_retries=retry))
        self.session.headers.update(
            {
                "User-Agent": "Mozilla/5.0 (compatible; UPrintingVariationExporter/1.0)",
                "Accept": "application/json, text/plain, */*",
            }
        )
        self.product_id = ""
        self.api_url = ""
        self.auth = ""
        self.defaults: dict[str, str] = {}
        self.visible_attr_ids: list[str] = []
        self.catalog: dict[str, Any] = {}
        self.product_image = ""

    def load(self) -> None:
        LOG.info("Product page download ho raha hai: %s", self.url)
        response = self.session.get(self.url, timeout=self.timeout)
        response.raise_for_status()
        html = response.text
        self.product_id = _first(r"var\s+page_product_id\s*=\s*[\"']?(\d+)", html, "product ID")
        if self.product_id == "0":
            raise ScraperError("Yeh category/landing page hai, configurable product page nahi. Specific product URL dein.")
        image_match = re.search(r"var\s+thumbnail_image\s*=\s*image_domain\s*\+\s*['\"]\/['\"]\s*\+\s*['\"]([^'\"]+)", html)
        image_domain_match = re.search(r"var\s+image_domain\s*=\s*['\"]([^'\"]+)", html)
        if image_match and image_domain_match:
            self.product_image = image_domain_match.group(1).rstrip("/") + "/" + image_match.group(1).lstrip("/")
        else:
            og = re.search(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)', html, re.I)
            self.product_image = og.group(1) if og else ""
        self.api_url = _first(
            r"api_compute_price_url\s*:\s*[\"'](https?:\\?/\\?/[^\"']+/v1)/computePrice",
            html,
            "calculator API URL",
        ).replace("\\/", "/")
        key = _first(r"clients\s*:\s*\{\s*key\s*:\s*[\"']([^\"']+)", html, "API client key")
        secret = _first(r"secret\s*:\s*[\"']([^\"']+)", html, "API client secret")
        self.auth = "Basic " + base64.b64encode(f"{key}:{secret}".encode()).decode()

        pricing_match = re.search(r"var\s+CalcPricingData\s*=\s*(\{.*?\});\s*window\.addEventListener", html, re.S)
        if pricing_match:
            pricing = json.loads(pricing_match.group(1))
            initial = pricing.get("request", {}).get("initial_price_data", {})
            normalized = pricing.get("response", {}).get("price", {}).get("price_data", {})
            if normalized:
                initial = {**initial, **{k: v for k, v in normalized.items() if str(k).startswith("attr")}}
        else:
            override_match = re.search(r'"price_data_override"\s*:\s*(\{[^}]*\})', html)
            initial = json.loads(override_match.group(1)) if override_match else {}
        self.defaults = {str(k): str(v) for k, v in initial.items() if str(k).startswith("attr")}

        visible_match = re.search(r'"visible_attrs"\s*:\s*\{\s*"attr_ids"\s*:\s*(\[[^]]*])', html)
        self.visible_attr_ids = [str(x) for x in json.loads(visible_match.group(1))] if visible_match else []
        self.catalog = self._post(f"getData/{self.product_id}", self._base_payload(include_product=False))
        # Page HTML can contain retired hidden option IDs. Keep overrides only
        # when they still exist in the current calculator catalog.
        clean_defaults: dict[str, str] = {}
        for attr_id, attr in self.catalog.get("prod_attrs", {}).items():
            key = f"attr{attr_id}"
            candidate = self.defaults.get(key)
            values = attr.get("prod_attr_vals", {})
            if candidate not in values:
                candidate = str(attr.get("default_value", ""))
            if candidate and candidate in values:
                clean_defaults[key] = candidate
        self.defaults = clean_defaults

    def _base_payload(self, include_product: bool = True) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "productType": "offset",
            "publishedVersion": True,
            "disableDataCache": False,
            "disablePriceCache": False,
        }
        if include_product:
            payload["product_id"] = self.product_id
        return payload

    def _post(self, endpoint: str, payload: dict[str, Any]) -> dict[str, Any]:
        headers = {"Authorization": self.auth, "Origin": "https://www.uprinting.com"}
        response = self.session.post(f"{self.api_url}/{endpoint}", json=payload, headers=headers, timeout=self.timeout)
        if response.status_code >= 400:
            detail = response.text[:500].replace("\n", " ")
            raise ScraperError(f"API {response.status_code}: {detail}")
        return response.json()

    def attributes(self, visible_only: bool = True) -> list[dict[str, Any]]:
        result = []
        visible = set(self.visible_attr_ids)
        for attr_id, attr in self.catalog.get("prod_attrs", {}).items():
            if visible_only and visible and str(attr_id) not in visible:
                continue
            if attr.get("hide_attribute_flag") == "y":
                continue
            values = []
            for value_id, value in attr.get("prod_attr_vals", {}).items():
                if value.get("hide_attribute_value_flag") == "y" or value.get("custom_flag") == "y":
                    continue
                # Dynamic-size products expose an internal max-dimension sentinel
                # alongside the customer-facing "Custom" option. The website hides it.
                factors = value.get("factors", {})
                if (
                    str(attr_id) == "3"
                    and self.catalog.get("dynamic_size") == "c"
                    and str(factors.get("width")) == str(self.catalog.get("end_width"))
                    and str(factors.get("height")) == str(self.catalog.get("end_height"))
                ):
                    continue
                values.append(
                    {
                        "option_id": str(value_id),
                        "source_attr_value_id": str(value.get("attr_val_id", "")),
                        "label": value.get("attr_value", ""),
                        "default": self.defaults.get(f"attr{attr_id}") == str(value_id),
                        "sort_order": value.get("sort_order"),
                        "factors": value.get("factors", {}),
                    }
                )
            values.sort(key=lambda v: (v["sort_order"] is None, v["sort_order"] or 0, v["label"]))
            result.append(
                {
                    "attribute_id": str(attr_id),
                    "name": attr.get("product_attribute_name") or attr.get("attribute_name") or f"Attribute {attr_id}",
                    "code": attr.get("attribute_code", ""),
                    "field_type": attr.get("field_type", ""),
                    "default_option_id": str(attr.get("default_value", "")),
                    "sort_order": attr.get("sort_order"),
                    "options": values,
                }
            )
        result.sort(key=lambda a: (a["sort_order"] is None, a["sort_order"] or 0))
        return result

    def price(self, selection: dict[str, str]) -> dict[str, Any]:
        payload = self._base_payload()
        payload.update(selection)
        return self._post("computePrice", payload)

    def _priced_row(self, selection: dict[str, str], changed: str = "") -> dict[str, Any]:
        response = self.price(selection)
        display = {
            str(item["attribute_id"]): {
                "name": item.get("attribute_name", ""),
                "label": item.get("attr_value", ""),
                "option_id": str(item.get("prod_attr_val_id", "")),
            }
            for item in response.get("display_specs", [])
        }
        return {
            "changed_attribute_id": changed,
            "selection": dict(selection),
            "display": display,
            "price": response.get("price"),
            "original_price": response.get("orig_price"),
            "total_price": response.get("total_price"),
            "unit_price": response.get("unit_price"),
            "quantity": response.get("qty"),
            "turnaround_days": response.get("turnaround"),
            "in_stock": response.get("in_stock_flag"),
            "currency": "USD",
            "order_specs": response.get("order_specs", []),
        }

    def selections(self, mode: str, vary: list[str] | None, max_combinations: int) -> Iterable[tuple[dict[str, str], str]]:
        attrs = self.attributes()
        by_id = {a["attribute_id"]: a for a in attrs}
        if mode == "default":
            yield dict(self.defaults), ""
            return
        if mode == "sweep":
            yield dict(self.defaults), ""
            seen = {tuple(sorted(self.defaults.items()))}
            for attr in attrs:
                for option in attr["options"]:
                    selection = dict(self.defaults)
                    selection[f"attr{attr['attribute_id']}"] = option["option_id"]
                    key = tuple(sorted(selection.items()))
                    if key not in seen:
                        seen.add(key)
                        yield selection, attr["attribute_id"]
            return

        chosen = list(by_id) if not vary or vary == ["all"] else vary
        unknown = [x for x in chosen if x not in by_id]
        if unknown:
            raise ValueError(f"Unknown --vary attribute IDs: {', '.join(unknown)}")
        option_lists = [[o["option_id"] for o in by_id[attr_id]["options"]] for attr_id in chosen]
        count = 1
        for options in option_lists:
            count *= len(options)
        if max_combinations and count > max_combinations:
            raise ValueError(
                f"{count:,} combinations banti hain, limit {max_combinations:,} hai. "
                "--vary mein kam attribute IDs dein ya --max-combinations barhayein (0 = unlimited)."
            )
        for values in itertools.product(*option_lists):
            selection = dict(self.defaults)
            selection.update({f"attr{attr_id}": option_id for attr_id, option_id in zip(chosen, values)})
            yield selection, ",".join(chosen)

    def scrape_prices(
        self,
        mode: str,
        vary: list[str] | None,
        max_combinations: int,
        workers: int,
        delay: float,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        selections = list(self.selections(mode, vary, max_combinations))
        LOG.info("%d price configurations process hongi", len(selections))
        prices: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []

        def run(item: tuple[dict[str, str], str]) -> dict[str, Any]:
            if delay:
                time.sleep(delay)
            return self._priced_row(*item)

        with ThreadPoolExecutor(max_workers=max(1, workers)) as pool:
            futures = {pool.submit(run, item): item for item in selections}
            for index, future in enumerate(as_completed(futures), 1):
                selection, changed = futures[future]
                try:
                    prices.append(future.result())
                except Exception as exc:  # invalid combos are expected in exhaustive mode
                    errors.append({"changed_attribute_id": changed, "selection": selection, "error": str(exc)})
                if index % 50 == 0 or index == len(selections):
                    LOG.info("Progress: %d/%d (valid=%d, invalid=%d)", index, len(selections), len(prices), len(errors))
        prices.sort(key=lambda row: tuple(sorted(row["selection"].items())))
        return prices, errors


def build_export(scraper: UPrintingScraper, mode: str, prices: list[dict[str, Any]], errors: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "metadata": {
            "source_url": scraper.url,
            "scraped_at_utc": datetime.now(timezone.utc).isoformat(),
            "product_id": scraper.product_id,
            "product_code": scraper.catalog.get("product_code"),
            "product_name": scraper.catalog.get("product_name"),
            "mode": mode,
            "currency": "USD",
            "valid_price_rows": len(prices),
            "invalid_rows": len(errors),
        },
        "default_selection": scraper.defaults,
        "attributes": scraper.attributes(),
        "prices": prices,
        "errors": errors,
    }


def save_json(data: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def _style_sheet(ws: Any) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = min(60, max(11, *(len(str(c.value or "")) + 2 for c in column[:300])))
        ws.column_dimensions[letter].width = width


def save_xlsx(data: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Field", "Value"])
    for key, value in data["metadata"].items():
        summary.append([key, value])
    summary.append(["default_selection", json.dumps(data["default_selection"], ensure_ascii=False)])

    attributes = wb.create_sheet("Attributes")
    attributes.append(["attribute_id", "attribute_name", "attribute_code", "option_id", "source_attr_value_id", "option_label", "is_default", "factors"])
    for attr in data["attributes"]:
        for option in attr["options"]:
            attributes.append([
                attr["attribute_id"], attr["name"], attr["code"], option["option_id"], option["source_attr_value_id"],
                option["label"], option["default"], json.dumps(option["factors"], ensure_ascii=False),
            ])

    attr_ids = [a["attribute_id"] for a in data["attributes"]]
    attr_names = {a["attribute_id"]: a["name"] for a in data["attributes"]}
    prices_ws = wb.create_sheet("Prices")
    fixed_headers = ["price", "original_price", "total_price", "unit_price", "currency", "quantity", "turnaround_days", "in_stock"]
    variation_headers = list(itertools.chain.from_iterable((f"{attr_names[x]} ID", attr_names[x]) for x in attr_ids))
    prices_ws.append(fixed_headers + variation_headers)
    for row in data["prices"]:
        cells = [row.get(x) for x in fixed_headers]
        for attr_id in attr_ids:
            shown = row.get("display", {}).get(attr_id, {})
            cells.extend([shown.get("option_id", row["selection"].get(f"attr{attr_id}", "")), shown.get("label", "")])
        prices_ws.append(cells)

    errors_ws = wb.create_sheet("Errors")
    errors_ws.append(["changed_attribute_id", "selection", "error"])
    for row in data["errors"]:
        errors_ws.append([row["changed_attribute_id"], json.dumps(row["selection"]), row["error"]])

    for ws in wb.worksheets:
        _style_sheet(ws)
    wb.save(path)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="UPrinting variation IDs aur live prices JSON/XLSX mein export karein.")
    parser.add_argument("url", nargs="?", default=DEFAULT_URL, help="UPrinting product URL")
    parser.add_argument("--mode", choices=("default", "sweep", "exhaustive"), default="sweep")
    parser.add_argument("--vary", default="", help="Exhaustive mode: comma-separated attribute IDs, ya 'all'")
    parser.add_argument("--max-combinations", type=int, default=10000, help="Safety cap; 0 = unlimited")
    parser.add_argument("--workers", type=int, default=4, help="Concurrent API calls (recommended <= 4)")
    parser.add_argument("--delay", type=float, default=0.05, help="Har request se pehle delay seconds")
    parser.add_argument("--output", default="uprinting_export", help="Output path without extension")
    parser.add_argument("--timeout", type=float, default=30)
    parser.add_argument("--verbose", action="store_true")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO, format="%(levelname)s: %(message)s")
    try:
        scraper = UPrintingScraper(args.url, timeout=args.timeout)
        scraper.load()
        vary = [x.strip() for x in args.vary.split(",") if x.strip()] or None
        prices, errors = scraper.scrape_prices(args.mode, vary, args.max_combinations, args.workers, args.delay)
        data = build_export(scraper, args.mode, prices, errors)
        base = Path(args.output)
        json_path, xlsx_path = base.with_suffix(".json"), base.with_suffix(".xlsx")
        save_json(data, json_path)
        save_xlsx(data, xlsx_path)
        print(f"Done: {len(prices)} valid prices, {len(errors)} invalid combinations")
        print(f"JSON : {json_path.resolve()}")
        print(f"Excel: {xlsx_path.resolve()}")
        return 0 if prices else 2
    except (requests.RequestException, ScraperError, ValueError, json.JSONDecodeError) as exc:
        LOG.error("%s", exc)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
